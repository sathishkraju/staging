from flask import *
from flask import Flask, redirect, url_for
from openpyxl import load_workbook
import requests
import json
import sys
sys.path.append('apicommon')
import apicommon
import base64
import xlsxwriter

app = Flask(__name__) #creating the Flask class object   
app=Flask(__name__,template_folder='templates')
 
@app.route('/') #decorator drfines the   
def index():
    book = load_workbook("account_group.xlsx")
    sheet = book.active
    return render_template('login.html',sheet=sheet)

@app.route('/generate_group/') #decorator drfines the   
def generate_group():
    userpass = apicommon.clientId+":"+apicommon.skey
    auth_access=base64.b64encode(userpass.encode()).decode()

    payload='grant_type=client_credentials'
    headers = {
      'Authorization': "Basic :"+auth_access,
      'Content-Type': 'application/x-www-form-urlencoded',
    }
    prnt=""
    response = requests.request("POST", apicommon.aurl, headers=headers, data=payload)
    access_token = (json.loads(response.content)["access_token"])
    headers = {
        'Accept': 'application/json',
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
        'X-BT-Pagination-Total':''
    }
    bearer="Bearer "+access_token
    com_file='account_group.xlsx';
    xlsx_File = xlsxwriter.Workbook(com_file)
    bold = xlsx_File.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'color':'white',
        'fg_color': '#1e4f87'})
    merge_format = xlsx_File.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'color':'white',
        'fg_color': '#879b20'})

    payload = "{subject:TestSubject,description:TestDescription,priority:Very Low}"
    headers = {
      'Authorization': bearer,
      'Content-Type': 'application/json',
      'Accept': 'application/json'
    }
    pNo=0
    exNo=0
    perPage=100
    n=0
    sheet_schedule = xlsx_File.add_worksheet("Account Group")
    sheet_schedule.write(0,0,"Id",bold)
    sheet_schedule.write(0,1,"Name",bold)
    for number in range(0, 100):
        if exNo==0:
            pNo = pNo+1
            url = apicommon.epurl+"/api/config/v1/vault/account-group?per_page="+str(perPage)+"&current_page="+str(pNo)        
            cl_resp = requests.request("GET", url, headers=headers, data=payload)
            resp = json.loads(cl_resp.text)       
            tot_len=len(resp)        
            if tot_len==perPage:
                for i in resp:
                    n=n+1                
                    sheet_schedule.write(n,0,str(i["id"]))
                    sheet_schedule.write(n,1,str(i["name"]))
            elif tot_len<perPage:
                if exNo==0:
                    for i in resp:
                        n=n+1
                        sheet_schedule.write(n,0,str(i["id"]))
                        sheet_schedule.write(n,1,str(i["name"]))
                    exNo=1
                    tot_len=0
                else:
                    exit
    xlsx_File.close()
    return "success"

@app.route('/account_group/',methods=['GET', 'POST'])  
def home():
    ulstr={}
    gid=""
    glid=""
    passwd=""
    status1=""
    status2=""
    status1_msg=""
    status2_msg=""
    print(request.method)
    if request.method == 'POST':
        gid = request.form.get('group_account')
        glid = request.form.get('group_account_list')
        passwd = request.form.get('pass')
        action = request.form.get('action')
            
    if gid!="":        
        userpass = apicommon.clientId+":"+apicommon.skey
        auth_access=base64.b64encode(userpass.encode()).decode()
        payload='grant_type=client_credentials'
        headers = {
        'Authorization': "Basic :"+auth_access,
        'Content-Type': 'application/x-www-form-urlencoded',
        }
        response = requests.request("POST", apicommon.aurl, headers=headers, data=payload)
        access_token = (json.loads(response.content)["access_token"])
        headers = {
        'Accept': 'application/json',
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
        'X-BT-Pagination-Total':''
        }
        bearer="Bearer "+access_token
        pNo=0
        exNo=0
        perPage=100
        n=0
        alist={}
        
        for number in range(0, 100):
            if exNo==0:
                pNo = pNo+1
                url = apicommon.epurl+"/api/config/v1/vault/account?account_group_id="+str(gid)+"&per_page="+str(perPage)+"&current_page="+str(pNo)
                cl_resp = requests.request("GET", url, headers=headers, data=payload)
                resp = json.loads(cl_resp.text)       
                tot_len=len(resp)        
                if tot_len==perPage:
                    for i in resp:
                        n=n+1                       
                        alist[str(i["id"])] = str(i["name"])         
                elif tot_len<perPage:
                    if exNo==0:
                        for i in resp:
                            n=n+1
                            alist[str(i["id"])] = str(i["name"])
                        exNo=1
                        tot_len=0
                    else:
                        exit
        ulstr=dict(sorted(alist.items(), key=lambda item: item[1].casefold()))
        print(action)
        if action!="":
            urlc = "https://nwncarousel.beyondtrustcloud.com/api/config/v1/vault/account/"+glid+"/force-check-in"
            payloadc = ""
            headersc = {
              'Content-Type': 'application/json',
              'Authorization': f'Bearer {access_token}',
            }
            responsec = requests.request("POST", urlc, headers=headersc, data=payloadc)
            respc=json.loads(responsec.text)
            print(responsec.status_code)
            if responsec.status_code!=200:
                status1=respc.get('message')           
            else:
                status1="Success"
            status1_msg="Force Check in Status : "+str(status1)
           
            if passwd!="":
                url1 = "https://nwncarousel.beyondtrustcloud.com/api/config/v1/vault/account/"+glid
                payload1 = json.dumps({
                  "password": passwd,
                  "private_key": passwd
                })
                headers1 = {
                  'Content-Type': 'application/json',
                  'Authorization': f'Bearer {access_token}',
                }
                response = requests.request("PATCH", url1, headers=headers1, data=payload1)
                respc1=json.loads(response.text)
                print(str("---")+str(response.status_code)+str("-------"))
                if response.status_code!=200:
                    status2=respc1.get('message')
                else:
                    status2="Success"
                status2_msg="Password Reset Status : "+str(status2)
    book = load_workbook("account_group.xlsx")
    sheet = book.active
    return render_template('login.html',sheet=sheet,len = len(ulstr), ulstr = ulstr,gid=gid,glid=glid,status1=status1,status2=status2,status1_msg=status1_msg,status2_msg=status2_msg)   

@app.route('/login',methods = ['POST'])  
def login():  
      uname=request.form['uname']  
      passwrd=request.form['pass']  
      if uname=="ayush" and passwrd=="google":  
          return "Welcome %s" %uname  
    
if __name__ =='__main__':  
    app.run(debug = True)